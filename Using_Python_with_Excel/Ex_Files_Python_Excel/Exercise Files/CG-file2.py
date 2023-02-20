from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('regions.xlsx')
ws = wb.active

cell_range = ws['A1':'C1']

col_range = ws['A':'C']
print(col_range)

row_range = ws[1:5]